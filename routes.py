from flask import render_template, request, redirect, url_for, flash, session, jsonify, make_response
from app import app, db
from models import User, WorkCenter, ProductionOrder, Department
from datetime import datetime, timedelta
import io
from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore

@app.route('/')
def login():
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def do_login():
    username = request.form['username']
    password = request.form['password']
    
    user = User.query.filter_by(username=username, is_active=True).first()
    
    if user and user.check_password(password):
        session['user_id'] = user.id
        session['username'] = user.username
        session['is_admin'] = user.is_admin
        
        if user.is_admin:
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('menu'))
    else:
        flash('Invalid username or password', 'error')
        return redirect(url_for('login'))

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/menu')
def menu():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('menu.html')

@app.route('/in_orders')
def in_orders():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get current user
    current_user = User.query.get(session['user_id'])
    
    # Filter work centers based on user's department
    if current_user and current_user.department:
        # Get departments that match user's department
        user_departments = Department.query.filter_by(name=current_user.department).all()
        if user_departments:
            # Get work centers assigned to user's department
            workcenters = WorkCenter.query.filter(
                WorkCenter.is_active == True,
                WorkCenter.departments.any(Department.id.in_([dept.id for dept in user_departments]))
            ).all()
        else:
            workcenters = []
    else:
        # If user has no department or admin, show all work centers
        workcenters = WorkCenter.query.filter_by(is_active=True).all()
    
    return render_template('in_orders.html', workcenters=workcenters)

@app.route('/out_orders')
def out_orders():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Get current user
    current_user = User.query.get(session['user_id'])
    
    # Filter work centers based on user's department
    if current_user and current_user.department:
        # Get departments that match user's department
        user_departments = Department.query.filter_by(name=current_user.department).all()
        if user_departments:
            # Get work centers assigned to user's department
            workcenters = WorkCenter.query.filter(
                WorkCenter.is_active == True,
                WorkCenter.departments.any(Department.id.in_([dept.id for dept in user_departments]))
            ).all()
        else:
            workcenters = []
    else:
        # If user has no department or admin, show all work centers
        workcenters = WorkCenter.query.filter_by(is_active=True).all()
    
    return render_template('out_orders.html', workcenters=workcenters)

@app.route('/save_orders', methods=['POST'])
def save_orders():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    order_type = request.form['order_type']
    orders_data = request.form.getlist('orders')
    
    try:
        for order_data in orders_data:
            if order_data:  # Skip empty entries
                parts = order_data.split('|')
                if len(parts) == 4:
                    workcenter_id, production_order, quantity, remark = parts
                    
                    # Skip if workcenter_id is empty or invalid
                    if not workcenter_id or workcenter_id == "":
                        continue
                    
                    new_order = ProductionOrder()
                    new_order.production_order = production_order.strip() if production_order else ""
                    new_order.workcenter_id = int(workcenter_id)
                    new_order.quantity = int(quantity) if quantity else 0
                    new_order.remark = remark.strip() if remark else ""
                    new_order.order_type = order_type
                    new_order.user_id = session['user_id']
                    db.session.add(new_order)
        
        db.session.commit()
        flash(f'{order_type} orders saved successfully!', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving orders: {str(e)}', 'error')
    
    return redirect(url_for('menu'))

@app.route('/reports')
def reports():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    search = request.args.get('search', '')
    workcenter_filter = request.args.get('workcenter', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = db.session.query(ProductionOrder).join(WorkCenter).join(User)
    
    # Apply filters
    if search:
        query = query.filter(ProductionOrder.production_order.contains(search))
    
    if workcenter_filter:
        query = query.filter(ProductionOrder.workcenter_id == int(workcenter_filter))
    
    if date_from:
        from datetime import datetime
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) >= date_from_obj)
    
    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) <= date_to_obj)
    
    # Default sorting by created_at desc
    orders = query.order_by(ProductionOrder.created_at.desc()).all()
    
    # Convert UTC times to IST for each order
    for order in orders:
        ist_time = order.created_at + timedelta(hours=5, minutes=30)
        order.created_at_ist = ist_time.strftime('%Y-%m-%d %H:%M:%S')
    
    # Get all work centers for the filter dropdown
    workcenters = WorkCenter.query.filter_by(is_active=True).all()
    
    # Get current user for Excel access check
    current_user = User.query.get(session['user_id'])
    has_excel_access = current_user and (current_user.excel_access or current_user.is_admin)
    
    return render_template('reports.html', orders=orders, search=search, 
                         workcenters=workcenters, workcenter_filter=workcenter_filter,
                         date_from=date_from, date_to=date_to, has_excel_access=has_excel_access)

@app.route('/balance_report')
def balance_report():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    search = request.args.get('search', '')
    workcenter_filter = request.args.get('workcenter', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get all production orders with user information
    query = db.session.query(ProductionOrder).join(WorkCenter).join(User)
    
    # Apply filters
    if search:
        query = query.filter(ProductionOrder.production_order.contains(search))
    
    if workcenter_filter:
        query = query.filter(ProductionOrder.workcenter_id == int(workcenter_filter))
    
    if date_from:
        from datetime import datetime
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) >= date_from_obj)
    
    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) <= date_to_obj)
    
    all_orders = query.all()
    
    # Calculate balance for each production order per work center
    balance_data = {}
    
    for order in all_orders:
        key = f"{order.production_order}_{order.workcenter_id}"
        
        if key not in balance_data:
            balance_data[key] = {
                'production_order': order.production_order,
                'workcenter_name': order.workcenter.name,
                'workcenter_id': order.workcenter_id,
                'user_name': order.user.name or order.user.username,
                'user_department': order.user.department or '-',
                'total_in': 0,
                'total_out': 0,
                'balance': 0,
                'remarks': set()  # Use set to avoid duplicate remarks
            }
        
        if order.order_type == 'IN':
            balance_data[key]['total_in'] += order.quantity
        else:
            balance_data[key]['total_out'] += order.quantity
        
        # Add remark if it exists and is not empty
        if order.remark and order.remark.strip():
            balance_data[key]['remarks'].add(order.remark.strip())
    
    # Calculate balance for each entry and convert remarks set to string
    for key in balance_data:
        balance_data[key]['balance'] = balance_data[key]['total_in'] - balance_data[key]['total_out']
        # Convert remarks set to comma-separated string
        balance_data[key]['remarks_text'] = ', '.join(sorted(balance_data[key]['remarks'])) if balance_data[key]['remarks'] else '-'
    
    # Convert to list and sort by production order, work center
    balance_list = list(balance_data.values())
    balance_list.sort(key=lambda x: (x['production_order'], x['workcenter_name']))
    
    # Get all work centers for the filter dropdown
    workcenters = WorkCenter.query.filter_by(is_active=True).all()
    
    # Get current user for Excel access check
    current_user = User.query.get(session['user_id'])
    has_excel_access = current_user and (current_user.excel_access or current_user.is_admin)
    
    return render_template('balance_report.html', balance_data=balance_list, search=search,
                         workcenters=workcenters, workcenter_filter=workcenter_filter,
                         date_from=date_from, date_to=date_to, has_excel_access=has_excel_access)

# Admin Routes
@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    # Dashboard statistics
    total_users = User.query.filter_by(is_active=True).count()
    total_workcenters = WorkCenter.query.filter_by(is_active=True).count()
    total_in_orders = ProductionOrder.query.filter_by(order_type='IN').count()
    total_out_orders = ProductionOrder.query.filter_by(order_type='OUT').count()
    
    # Recent orders
    recent_orders = ProductionOrder.query.join(WorkCenter).join(User).order_by(ProductionOrder.created_at.desc()).limit(10).all()
    
    return render_template('admin_dashboard.html', 
                         total_users=total_users,
                         total_workcenters=total_workcenters,
                         total_in_orders=total_in_orders,
                         total_out_orders=total_out_orders,
                         recent_orders=recent_orders)

@app.route('/admin/users')
def admin_users():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    users = User.query.all()
    departments = Department.query.filter_by(is_active=True).all()
    return render_template('admin_users.html', users=users, departments=departments)

@app.route('/admin/create_user', methods=['POST'])
def create_user():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    username = request.form['username']
    name = request.form.get('name', '')
    department = request.form.get('department', '')
    password = request.form['password']
    is_admin = 'is_admin' in request.form
    excel_access = 'excel_access' in request.form
    
    # Check if username already exists
    existing_user = User.query.filter_by(username=username).first()
    if existing_user:
        flash('Username already exists', 'error')
        return redirect(url_for('admin_users'))
    
    try:
        new_user = User()
        new_user.username = username
        new_user.name = name
        new_user.department = department
        new_user.is_admin = is_admin
        new_user.excel_access = excel_access
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        flash('User created successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating user: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/edit_user/<int:user_id>', methods=['POST'])
def edit_user(user_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(user_id)
    new_username = request.form['username']
    
    # Check if username already exists (excluding current user)
    existing_user = User.query.filter_by(username=new_username).first()
    if existing_user and existing_user.id != user_id:
        flash('Username already exists', 'error')
        return redirect(url_for('admin_users'))
    
    user.username = new_username
    user.name = request.form.get('name', '')
    user.department = request.form.get('department', '')
    password = request.form.get('password', '').strip()
    if password:
        user.set_password(password)
    user.is_admin = 'is_admin' in request.form
    user.excel_access = 'excel_access' in request.form
    user.is_active = 'is_active' in request.form
    
    try:
        db.session.commit()
        flash('User updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating user: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/delete_user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting the current admin user
    if user.id == session['user_id']:
        flash('Cannot delete your own account', 'error')
        return redirect(url_for('admin_users'))
    
    try:
        # Delete user completely from database
        db.session.delete(user)
        db.session.commit()
        flash('User deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('admin_users'))

@app.route('/admin/reports')
def admin_reports():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    search = request.args.get('search', '')
    workcenter_filter = request.args.get('workcenter', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = db.session.query(ProductionOrder).join(WorkCenter).join(User)
    
    # Apply filters
    if search:
        query = query.filter(ProductionOrder.production_order.contains(search))
    
    if workcenter_filter:
        query = query.filter(ProductionOrder.workcenter_id == int(workcenter_filter))
    
    if date_from:
        from datetime import datetime
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) >= date_from_obj)
    
    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) <= date_to_obj)
    
    # Default sorting by created_at desc
    orders = query.order_by(ProductionOrder.created_at.desc()).all()
    
    # Convert UTC times to IST for each order
    for order in orders:
        ist_time = order.created_at + timedelta(hours=5, minutes=30)
        order.created_at_ist = ist_time.strftime('%Y-%m-%d %H:%M:%S')
    
    # Get all work centers for the filter dropdown
    workcenters = WorkCenter.query.filter_by(is_active=True).all()
    
    return render_template('admin_reports.html', orders=orders, search=search,
                         workcenters=workcenters, workcenter_filter=workcenter_filter,
                         date_from=date_from, date_to=date_to)

@app.route('/admin/balance_report')
def admin_balance_report():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    search = request.args.get('search', '')
    workcenter_filter = request.args.get('workcenter', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Get all production orders with user information
    query = db.session.query(ProductionOrder).join(WorkCenter).join(User)
    
    # Apply filters
    if search:
        query = query.filter(ProductionOrder.production_order.contains(search))
    
    if workcenter_filter:
        query = query.filter(ProductionOrder.workcenter_id == int(workcenter_filter))
    
    if date_from:
        from datetime import datetime
        date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) >= date_from_obj)
    
    if date_to:
        from datetime import datetime
        date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
        query = query.filter(db.func.date(ProductionOrder.created_at) <= date_to_obj)
    
    all_orders = query.all()
    
    # Calculate balance for each production order per work center per user
    balance_data = {}
    
    for order in all_orders:
        key = f"{order.production_order}_{order.workcenter_id}_{order.user_id}"
        
        if key not in balance_data:
            balance_data[key] = {
                'production_order': order.production_order,
                'workcenter_name': order.workcenter.name,
                'workcenter_id': order.workcenter_id,
                'user_name': order.user.name or order.user.username,
                'user_department': order.user.department or '-',
                'total_in': 0,
                'total_out': 0,
                'balance': 0,
                'last_activity': order.created_at,
                'remarks': set()  # Use set to avoid duplicate remarks
            }
        
        if order.order_type == 'IN':
            balance_data[key]['total_in'] += order.quantity
        else:
            balance_data[key]['total_out'] += order.quantity
        
        # Track the latest activity date
        if order.created_at > balance_data[key]['last_activity']:
            balance_data[key]['last_activity'] = order.created_at
        
        # Add remark if it exists and is not empty
        if order.remark and order.remark.strip():
            balance_data[key]['remarks'].add(order.remark.strip())
    
    # Calculate balance for each entry and convert remarks set to string
    for key in balance_data:
        balance_data[key]['balance'] = balance_data[key]['total_in'] - balance_data[key]['total_out']
        # Convert UTC time to IST (UTC + 5:30)
        ist_time = balance_data[key]['last_activity'] + timedelta(hours=5, minutes=30)
        balance_data[key]['last_activity_ist'] = ist_time.strftime('%Y-%m-%d %H:%M:%S')
        # Convert remarks set to comma-separated string
        balance_data[key]['remarks_text'] = ', '.join(sorted(balance_data[key]['remarks'])) if balance_data[key]['remarks'] else '-'
    
    # Convert to list and sort by production order, work center, then user name
    balance_list = list(balance_data.values())
    balance_list.sort(key=lambda x: (x['production_order'], x['workcenter_name'], x['user_name']))
    
    # Get all work centers for the filter dropdown
    workcenters = WorkCenter.query.filter_by(is_active=True).all()
    
    return render_template('admin_balance_report.html', balance_data=balance_list, search=search,
                         workcenters=workcenters, workcenter_filter=workcenter_filter,
                         date_from=date_from, date_to=date_to)

@app.route('/admin/export_excel')
def export_excel():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    # Create workbook with two worksheets
    wb = Workbook()
    
    # First worksheet: All Production Orders
    ws1 = wb.active
    if ws1 is not None:
        ws1.title = "All Production Orders"
    
    # Define headers for all orders
    headers1 = ['Production Order', 'Work Center', 'Quantity', 'Type', 'Remark', 'Name', 'Department', 'Date & Time']
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add headers for first sheet
    if ws1 is not None:
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)  # type: ignore
            if cell is not None:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        # Get all orders
        orders = db.session.query(ProductionOrder).join(WorkCenter).join(User).order_by(ProductionOrder.created_at.desc()).all()
        
        # Add data to first sheet
        for row, order in enumerate(orders, 2):
            ws1.cell(row=row, column=1, value=order.production_order)  # type: ignore
            ws1.cell(row=row, column=2, value=order.workcenter.name)  # type: ignore
            ws1.cell(row=row, column=3, value=order.quantity)  # type: ignore
            ws1.cell(row=row, column=4, value=order.order_type)  # type: ignore
            ws1.cell(row=row, column=5, value=order.remark or '-')  # type: ignore
            ws1.cell(row=row, column=6, value=order.user.name or order.user.username)  # type: ignore
            ws1.cell(row=row, column=7, value=order.user.department or '-')  # type: ignore
            # Convert to IST (UTC + 5:30) for display
            ist_time = order.created_at + timedelta(hours=5, minutes=30)
            ws1.cell(row=row, column=8, value=ist_time.strftime('%Y-%m-%d %H:%M:%S') + ' IST')  # type: ignore
    
    # Second worksheet: Balance Report
    ws2 = wb.create_sheet(title="Balance Report")
    
    # Define headers for balance report
    headers2 = ['Production Order', 'Work Center', 'Remarks', 'Total IN', 'Total OUT', 'Balance']
    
    # Add headers for second sheet
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)  # type: ignore
        if cell is not None:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    # Calculate balance data (same logic as balance_report route)
    balance_data = {}
    
    for order in orders:
        key = f"{order.production_order}_{order.workcenter_id}"
        
        if key not in balance_data:
            balance_data[key] = {
                'production_order': order.production_order,
                'workcenter_name': order.workcenter.name,
                'workcenter_id': order.workcenter_id,
                'total_in': 0,
                'total_out': 0,
                'balance': 0,
                'remarks': set()
            }
        
        if order.order_type == 'IN':
            balance_data[key]['total_in'] += order.quantity
        else:
            balance_data[key]['total_out'] += order.quantity
        
        # Add remark if it exists and is not empty
        if order.remark and order.remark.strip():
            balance_data[key]['remarks'].add(order.remark.strip())
    
    # Calculate balance for each entry and convert remarks set to string
    for key in balance_data:
        balance_data[key]['balance'] = balance_data[key]['total_in'] - balance_data[key]['total_out']
        # Convert remarks set to comma-separated string
        balance_data[key]['remarks_text'] = ', '.join(sorted(balance_data[key]['remarks'])) if balance_data[key]['remarks'] else '-'
    
    # Convert to list and sort by production order
    balance_list = list(balance_data.values())
    balance_list.sort(key=lambda x: (x['production_order'], x['workcenter_name']))
    
    # Add balance data to second sheet
    for row, item in enumerate(balance_list, 2):
        ws2.cell(row=row, column=1, value=item['production_order'])  # type: ignore
        ws2.cell(row=row, column=2, value=item['workcenter_name'])  # type: ignore
        ws2.cell(row=row, column=3, value=item['remarks_text'])  # type: ignore
        ws2.cell(row=row, column=4, value=item['total_in'])  # type: ignore
        ws2.cell(row=row, column=5, value=item['total_out'])  # type: ignore
        ws2.cell(row=row, column=6, value=item['balance'])  # type: ignore
    
    # Auto-adjust column widths for both sheets
    for ws in [ws1, ws2]:
        if ws and hasattr(ws, 'columns') and hasattr(ws, 'column_dimensions'):
            for column in ws.columns:  # type: ignore
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename=production_orders_with_balance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@app.route('/export_excel')
def user_export_excel():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    # Check if user has Excel access
    current_user = User.query.get(session['user_id'])
    if not current_user or (not current_user.excel_access and not current_user.is_admin):
        flash('Access denied. Excel export permission required.', 'error')
        return redirect(url_for('reports'))
    
    # Get current user for department filtering
    user_department = current_user.department
    
    # Create workbook with two worksheets
    wb = Workbook()
    
    # First worksheet: Production Orders (filtered by user's department if not admin)
    ws1 = wb.active
    if ws1 is not None:
        ws1.title = "Production Orders"
    
    # Define headers for orders
    headers1 = ['Production Order', 'Work Center', 'Quantity', 'Type', 'Remark', 'Name', 'Department', 'Date & Time']
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add headers for first sheet
    if ws1 is not None:
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)  # type: ignore
            if cell is not None:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        # Get orders filtered by user's department access (unless admin)
        query = db.session.query(ProductionOrder).join(WorkCenter).join(User)
        if not current_user.is_admin and user_department:
            query = query.filter(User.department == user_department)
        
        orders = query.order_by(ProductionOrder.created_at.desc()).all()
        
        # Add data to first sheet
        for row, order in enumerate(orders, 2):
            ws1.cell(row=row, column=1, value=order.production_order)  # type: ignore
            ws1.cell(row=row, column=2, value=order.workcenter.name)  # type: ignore
            ws1.cell(row=row, column=3, value=order.quantity)  # type: ignore
            ws1.cell(row=row, column=4, value=order.order_type)  # type: ignore
            ws1.cell(row=row, column=5, value=order.remark or '-')  # type: ignore
            ws1.cell(row=row, column=6, value=order.user.name or order.user.username)  # type: ignore
            ws1.cell(row=row, column=7, value=order.user.department or '-')  # type: ignore
            # Convert to IST (UTC + 5:30) for display
            ist_time = order.created_at + timedelta(hours=5, minutes=30)
            ws1.cell(row=row, column=8, value=ist_time.strftime('%Y-%m-%d %H:%M:%S') + ' IST')  # type: ignore
    
    # Second worksheet: Balance Report
    ws2 = wb.create_sheet(title="Balance Report")
    
    # Define headers for balance report
    headers2 = ['Production Order', 'Work Center', 'Remarks', 'Total IN', 'Total OUT', 'Balance']
    
    # Add headers for second sheet
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)  # type: ignore
        if cell is not None:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
    
    # Get balance data (filtered by user's department if not admin)
    query = db.session.query(ProductionOrder).join(WorkCenter).join(User)
    if not current_user.is_admin and user_department:
        query = query.filter(User.department == user_department)
    
    all_orders = query.all()
    
    # Calculate balance for each production order per work center
    balance_data = {}
    
    for order in all_orders:
        key = f"{order.production_order}_{order.workcenter_id}"
        
        if key not in balance_data:
            balance_data[key] = {
                'production_order': order.production_order,
                'workcenter_name': order.workcenter.name,
                'workcenter_id': order.workcenter_id,
                'total_in': 0,
                'total_out': 0,
                'balance': 0,
                'remarks': set()
            }
        
        if order.order_type == 'IN':
            balance_data[key]['total_in'] += order.quantity
        else:
            balance_data[key]['total_out'] += order.quantity
        
        # Add remark if it exists and is not empty
        if order.remark and order.remark.strip():
            balance_data[key]['remarks'].add(order.remark.strip())
    
    # Calculate balance for each entry and convert remarks set to string
    for key in balance_data:
        balance_data[key]['balance'] = balance_data[key]['total_in'] - balance_data[key]['total_out']
        # Convert remarks set to comma-separated string
        balance_data[key]['remarks_text'] = ', '.join(sorted(balance_data[key]['remarks'])) if balance_data[key]['remarks'] else '-'
    
    # Convert to list and sort by production order
    balance_list = list(balance_data.values())
    balance_list.sort(key=lambda x: (x['production_order'], x['workcenter_name']))
    
    # Add balance data to second sheet
    for row, item in enumerate(balance_list, 2):
        ws2.cell(row=row, column=1, value=item['production_order'])  # type: ignore
        ws2.cell(row=row, column=2, value=item['workcenter_name'])  # type: ignore
        ws2.cell(row=row, column=3, value=item['remarks_text'])  # type: ignore
        ws2.cell(row=row, column=4, value=item['total_in'])  # type: ignore
        ws2.cell(row=row, column=5, value=item['total_out'])  # type: ignore
        ws2.cell(row=row, column=6, value=item['balance'])  # type: ignore
    
    # Auto-adjust column widths for both sheets
    for ws in [ws1, ws2]:
        if ws and hasattr(ws, 'columns') and hasattr(ws, 'column_dimensions'):
            for column in ws.columns:  # type: ignore
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = make_response(output.read())
    response.headers['Content-Disposition'] = f'attachment; filename=production_orders_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    
    return response

@app.route('/admin/master_data')
def master_data():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    workcenters = WorkCenter.query.all()
    departments = Department.query.all()
    return render_template('master_data.html', workcenters=workcenters, departments=departments)

@app.route('/admin/create_workcenter', methods=['POST'])
def create_workcenter():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    name = request.form['name']
    department_ids = request.form.getlist('departments')
    
    try:
        new_workcenter = WorkCenter()
        new_workcenter.name = name
        
        # Assign departments to work center
        if department_ids:
            departments = Department.query.filter(Department.id.in_(department_ids)).all()
            new_workcenter.departments = departments
        
        db.session.add(new_workcenter)
        db.session.commit()
        flash('Work center created successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating work center: {str(e)}', 'error')
    
    return redirect(url_for('master_data'))

@app.route('/admin/edit_workcenter/<int:wc_id>', methods=['POST'])
def edit_workcenter(wc_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    workcenter = WorkCenter.query.get_or_404(wc_id)
    workcenter.name = request.form['name']
    workcenter.is_active = 'is_active' in request.form
    
    # Update department assignments
    department_ids = request.form.getlist('departments')
    if department_ids:
        departments = Department.query.filter(Department.id.in_(department_ids)).all()
        workcenter.departments = departments
    else:
        workcenter.departments = []
    
    try:
        db.session.commit()
        flash('Work center updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating work center: {str(e)}', 'error')
    
    return redirect(url_for('master_data'))

@app.route('/admin/delete_workcenter/<int:wc_id>', methods=['POST'])
def delete_workcenter(wc_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    workcenter = WorkCenter.query.get_or_404(wc_id)
    
    try:
        # Delete work center completely from database
        db.session.delete(workcenter)
        db.session.commit()
        flash('Work center deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting work center: {str(e)}', 'error')
    
    return redirect(url_for('master_data'))

# Department Management Routes
@app.route('/admin/create_department', methods=['POST'])
def create_department():
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    name = request.form['name']
    
    try:
        new_department = Department()
        new_department.name = name
        db.session.add(new_department)
        db.session.commit()
        flash('Department created successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error creating department: {str(e)}', 'error')
    
    return redirect(url_for('master_data'))

@app.route('/admin/edit_department/<int:dept_id>', methods=['POST'])
def edit_department(dept_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    department = Department.query.get_or_404(dept_id)
    department.name = request.form['name']
    department.is_active = 'is_active' in request.form
    
    try:
        db.session.commit()
        flash('Department updated successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating department: {str(e)}', 'error')
    
    return redirect(url_for('master_data'))

@app.route('/admin/delete_department/<int:dept_id>', methods=['POST'])
def delete_department(dept_id):
    if 'user_id' not in session or not session.get('is_admin'):
        return redirect(url_for('login'))
    
    department = Department.query.get_or_404(dept_id)
    
    try:
        # Delete department completely from database
        db.session.delete(department)
        db.session.commit()
        flash('Department deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting department: {str(e)}', 'error')
    
    return redirect(url_for('master_data'))

# Bulk Delete Routes
@app.route('/admin/bulk_delete_orders', methods=['POST'])
def bulk_delete_orders():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    order_ids = request.form.getlist('order_ids')
    
    if not order_ids:
        flash('No orders selected for deletion.', 'warning')
        return redirect(url_for('admin_reports'))
    
    try:
        # Convert to integers for safety
        order_ids = [int(order_id) for order_id in order_ids]
        
        # Delete selected orders
        deleted_count = ProductionOrder.query.filter(ProductionOrder.id.in_(order_ids)).delete()
        db.session.commit()
        
        flash(f'Successfully deleted {deleted_count} production order(s).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting orders: {str(e)}', 'error')
    
    return redirect(url_for('admin_reports'))

@app.route('/admin/bulk_delete_by_production_order', methods=['POST'])
def bulk_delete_by_production_order():
    if 'user_id' not in session or not session.get('is_admin'):
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('login'))
    
    production_orders = request.form.getlist('production_orders')
    
    if not production_orders:
        flash('No production orders selected for deletion.', 'warning')
        return redirect(url_for('balance_report'))
    
    try:
        deleted_count = 0
        
        for production_order_data in production_orders:
            # Extract production order and workcenter_id from the value
            parts = production_order_data.split('-')
            if len(parts) >= 2:
                production_order = '-'.join(parts[:-1])  # Handle production orders with dashes
                workcenter_id = int(parts[-1])
                
                # Delete all orders for this production order and work center
                count = ProductionOrder.query.filter_by(
                    production_order=production_order,
                    workcenter_id=workcenter_id
                ).delete()
                deleted_count += count
        
        db.session.commit()
        flash(f'Successfully deleted {deleted_count} production order(s).', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting production orders: {str(e)}', 'error')
    
    return redirect(url_for('balance_report'))
